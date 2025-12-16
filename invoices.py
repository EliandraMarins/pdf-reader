import os
from openpyxl import Workbook
import pdfplumber
import re
from datetime import datetime
import mysql.connector

def main():
    
    # Pegar a pasta em que os pds estao e fazer a contagem (maior que zero)
    try:
        directory = 'pdf_invoices'
        files = os.listdir(directory)
        files_quantity = len(files)
    except Exception as e:
        print(f"An error ocurred: {e}")


    if files_quantity == 0:
        raise Exception("No files found in the directory.")

    #Estrutura do arquivo excel - onde as informacoes serao padronizadas
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = 'Invoice Imports'

    #Estabelecer o nome das colunas - array
        ws['A1'] = 'Invoice #'
        ws['B1'] = 'Date'
        ws['C1'] = 'File Name'
        ws['D1'] = 'Status' #Sempre é pra existir
    

    #Checar se a linha esta vazia, senao, ir para a proxima
        last_empty_line = 1
        while ws['D' + str(last_empty_line)].value is not None:
            last_empty_line += 1

    except Exception as e:
        print(f"It was not possible to create excel file: {e}")         

    #Pegar os valores - Abre o arquivo pdf, pega todo o texto, extrai desse texto o numero da 
    # nota fiscal, data, status e coloca tudo isso no excel
    for file in files:
        try:
            with pdfplumber.open(directory + "/" + file) as pdf:
                first_page = pdf.pages[0]
                pdf_text = first_page.extract_text()


        #print(pdf_text)

        #Utilizar regex pra pegar o numero do invoice dentro de todo esse texto
        #Quando temos muito texto e queremos extrair algo especifico, eh possivel estabelecer um padrao
        #E atraves dessa sequencia ele entende que deve buscar valores baseados nos parametros passados

            inv_number_re_pattern = r'INVOICE #(\d+)'
            inv_date_re_pattern = r'DATE: (\d{2}/\d{2}/\d{4})'

        #Com a instrucao regex, procura e faz as verificacoes no excel caso eles forem encontrados

            match_number = re.search(inv_number_re_pattern, pdf_text)
            match_date = re.search(inv_date_re_pattern, pdf_text)

            if match_number:
                ws['A{}'.format(last_empty_line)] = match_number.group(1)
            else:
                raise Exception("Couldn't find invoice number.")     
                

            if match_date:
                ws['B{}'.format(last_empty_line)] = match_date.group(1)
            else:
                raise Exception("Couldn't find invoice date.")



            ws['C{}'.format(last_empty_line)] = file
            ws['D{}'.format(last_empty_line)] = "Completed"

            last_empty_line += 1

        except Exception as e:
            print(f"Error processing file: {e}")
            ws['D{}'.format(last_empty_line)] = "Exception{}".format(e) #Aparece couldn't find invoice number or date  
            
            last_empty_line += 1 #Deu exception e tem que ir para o prox item

    #Pegar data e horario pra nomear o arquivo excel
    try:
        full_now = str(datetime.now()).replace(":", "-")
        dot_index = full_now.index(".")
        now = full_now[:dot_index]

        #Salvar arquivo excel
        wb.save("Invoices - {}.xlsx".format(now))

    except Exception as e:
        print(f"It was not possible to save the excel file: {e}")


if __name__ == "__main__":
    main()